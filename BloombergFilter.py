import pandas as pd
import openpyxl
import re
import os

def clean_bloomberg_excel(file_path):
    """
    Cleans a Bloomberg-formatted Excel file and returns a data frame.
    Skips the initial header rows and only extracts date and PX_LAST columns.
    """
     # load excel file
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    nrows = sheet.max_row

    start_idx = None
    for i in range(1, nrows + 1):  # 1-indexed, so from 1 to max_row inclusive
        first_cell = sheet.cell(row=i, column=1).value
        if first_cell is None:
            continue
        first_cell_str = str(first_cell).strip().lower()
        if first_cell_str == "date":
            start_idx = i
            break

    if start_idx is None:
        raise ValueError(f"No 'Date' header found in {file_path}")


    df = pd.read_excel(file_path, skiprows=start_idx - 1, header=0) # skip all rows above header
    df.replace(",", ".", regex=True, inplace=True) # Replace commas so we can actually use in python

    # just a check to make sure the right columns exist
    if 'PX_LAST' not in df.columns or 'Date' not in df.columns:
        raise ValueError(f"'Date' and/or 'PX_LAST' columns not found in {file_path}")

    df = df[['Date', 'PX_LAST']]
    df['PX_LAST'] = df['PX_LAST'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'])


    # drop the most recent price, generally bugged out because it isn't the final price of the day, since extracted in the middle of it
    df = df.iloc[1:].reset_index(drop=True)

    return df


### Below is the script to clean all (EXCEL) files in a folder

raw_dir = "raw_Dat"         # Folder with raw Excel files
cleaned_dir = "clean_Dat"     # Folder to save cleaned Excel files
cleaned_data = {}

os.makedirs(cleaned_dir, exist_ok=True) # Make sure output folder exists, else creates

# Process each .xlsx file
for filename in os.listdir(raw_dir):
    if filename.endswith(".xlsx"):
        full_path = os.path.join(raw_dir, filename)
        try:
            df = clean_bloomberg_excel(full_path)
            cleaned_data[filename] = df

            # Save cleaned data
            target_path = os.path.join(cleaned_dir, filename)
            df.to_excel(target_path, index=False)

            print(f"Successful cleaning: {filename}")
        except Exception as e:
            print(f"Failed to clean {filename}: {e}")


# Combine all: Date | equity1 | equity2 | ...
combined_df = pd.DataFrame()

for filename, df in cleaned_data.items():
    temp_df = df[['Date', 'PX_LAST']].copy()
    equity_name = filename.replace('.xlsx', '') 
    temp_df.rename(columns={'PX_LAST': equity_name}, inplace=True)
    
    if combined_df.empty:
        combined_df = temp_df
    else:
        combined_df = pd.merge(combined_df, temp_df, on='Date', how='outer') # extracts the price, and joins on the outside to match the format

combined_file_path = os.path.join(cleaned_dir, 'Combined_data.xlsx')
combined_df.to_excel(combined_file_path, index=False)

print(f"Combined data saved as: {combined_file_path}")